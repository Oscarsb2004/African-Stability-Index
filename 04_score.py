"""
04_score.py — Aggregate normalized scores into pillar and overall index scores.

Input:  data/03_norm.xlsx  (norm_data sheet)
Output: data/04_scores.xlsx

Five scoring methods (all produce overall scores in [0, 100]):

  equal     — arithmetic mean of pillar scores, weights = 1/7 each.
              Simple, interpretable, fully compensatory.

  pca       — first principal component loadings as pillar weights.
              Data-driven; weights reflect variance structure.

  bod       — Benefit of Doubt (Cherchye et al. 2007): per-country DEA-CCR
              LP that maximises own composite score, subject to global
              weight bounds [weight_min, weight_max] per pillar.
              Requires: pip install pulp

  entropy   — Shannon entropy weights: indicators / pillars with high
              cross-country variance receive higher weight.
              Rewards informativeness; penalises zero-variance pillars.

  geometric — Geometric mean of pillar scores: exp(sum(w_i * ln(s_i))).
              Partial compensability — low performance on one pillar cannot
              be fully offset by excellence on another.
              Methodologically preferred for stability indices (avoids the
              full-compensability flaw of arithmetic aggregation).

Pillar scores (all methods):
  Arithmetic mean of within-pillar normalized indicator scores.
  Countries with all pillar indicators missing receive NaN pillar score.
  Countries with some indicators missing use the mean of available ones.

Run from the project root:
    python 04_score.py
"""

import sys
import logging
import time as _time
import warnings

import numpy as np
import pandas as pd
import yaml
from pathlib import Path
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

from constants import (
    PILLAR_DEFS, ACTIVE_PRESET, WEIGHT_PRESETS,
    WEIGHT_MIN, WEIGHT_MAX, SMALL,
)

# ── Logging ────────────────────────────────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format="%(levelname)-8s %(message)s",
    stream=sys.stdout,
)
logger = logging.getLogger(__name__)

# ── Paths ──────────────────────────────────────────────────────────────────────

INDICATORS_DIR = Path("indicators_list")
INPUT_FILE     = Path("data/03_norm.xlsx")
OUTPUT_DIR     = Path("data")
OUTPUT_FILE    = OUTPUT_DIR / "04_scores.xlsx"

# ── Config ─────────────────────────────────────────────────────────────────────

# PILLAR_DEFS, WEIGHT_PRESETS, ACTIVE_PRESET, WEIGHT_MIN, WEIGHT_MAX, SMALL
# are imported from constants.py — do not redefine them here.


# ── Indicator metadata ─────────────────────────────────────────────────────────

def load_pillar_map() -> dict[str, list[str]]:
    """Returns {pillar_id: [variable_name, ...]} from all indicator YAMLs."""
    pillar_map: dict[str, list[str]] = {k: [] for k in PILLAR_DEFS}
    for path in sorted(INDICATORS_DIR.glob("*.yaml")):
        with open(path, "r", encoding="utf-8") as f:
            data = yaml.safe_load(f)
        if not data:
            continue
        for ind in data:
            if ind.get("role", "scoring") != "scoring":
                continue
            for p in ind.get("pillars", []):
                if p in pillar_map and ind["variable_name"] not in pillar_map[p]:
                    pillar_map[p].append(ind["variable_name"])
    return pillar_map


# ── Step 1: Pillar scores ──────────────────────────────────────────────────────

def compute_pillar_scores(
    norm_wide: pd.DataFrame,
    pillar_map: dict[str, list[str]],
) -> pd.DataFrame:
    """
    norm_wide: rows=iso3, cols=variable_name (scores 0-100, NaN where missing).
    Returns: pillar_wide — rows=iso3, cols=pillar_id (0-100, NaN if all missing).
    """
    result = {}
    for pillar_id, vars_in_pillar in pillar_map.items():
        available = [v for v in vars_in_pillar if v in norm_wide.columns]
        if not available:
            logger.warning("Pillar %s: no indicators in norm_data — pillar will be NaN.", pillar_id)
            result[pillar_id] = pd.Series(float("nan"), index=norm_wide.index)
            continue
        sub = norm_wide[available]
        # Mean of available (non-NaN) indicators per country
        result[pillar_id] = sub.mean(axis=1, skipna=True)
        # Countries where ALL indicators are NaN → pillar is NaN
        all_nan_mask = sub.isna().all(axis=1)
        result[pillar_id][all_nan_mask] = float("nan")
        n_complete = (sub.notna().all(axis=1)).sum()
        logger.info(
            "  Pillar %s (%s): %d indicators, %d/%d countries fully complete",
            pillar_id, PILLAR_DEFS[pillar_id], len(available),
            n_complete, len(norm_wide),
        )
    return pd.DataFrame(result, index=norm_wide.index)


# ── Step 2: Scoring methods ────────────────────────────────────────────────────

def score_equal(pillar_wide: pd.DataFrame) -> pd.Series:
    """Weighted mean using ACTIVE_PRESET weights from WEIGHT_PRESETS."""
    pillars = [p for p in PILLAR_DEFS if p in pillar_wide.columns]
    preset  = WEIGHT_PRESETS[ACTIVE_PRESET]
    w_arr   = np.array([preset[p] for p in pillars])
    w_arr   = w_arr / w_arr.sum()   # renormalise in case preset doesn't cover all pillars

    scores = {}
    for iso3, row in pillar_wide[pillars].iterrows():
        vals  = row.values.astype(float)
        valid = ~np.isnan(vals)
        if valid.sum() == 0:
            scores[iso3] = float("nan")
        else:
            w_valid = w_arr[valid] / w_arr[valid].sum()
            scores[iso3] = float(np.dot(w_valid, vals[valid]))
    return pd.Series(scores)


def score_pca(pillar_wide: pd.DataFrame) -> pd.Series:
    """
    First PC loadings as pillar weights.
    Countries missing any pillar are excluded from PCA fit,
    but scores are predicted for all using fitted loadings.
    Returns scores scaled to [0, 100].
    """
    from sklearn.decomposition import PCA
    from sklearn.preprocessing import StandardScaler

    pillars = [p for p in PILLAR_DEFS if p in pillar_wide.columns]
    sub = pillar_wide[pillars]
    complete_mask = sub.notna().all(axis=1)

    if complete_mask.sum() < len(pillars):
        logger.warning("PCA: only %d/%d countries have complete pillar data.",
                       complete_mask.sum(), len(sub))

    X_fit = sub[complete_mask].values
    scaler = StandardScaler()
    X_scaled = scaler.fit_transform(X_fit)
    pca = PCA(n_components=1)
    pca.fit(X_scaled)

    loadings = pca.components_[0]
    # Orient PC1: if the majority of loadings are negative, flip the whole vector.
    # This ensures PC1 points in the "higher stability = higher score" direction.
    if (loadings < 0).sum() > len(loadings) / 2:
        loadings = -loadings
    # After orientation, any remaining negative loadings indicate pillars that load
    # INVERSELY on the stability dimension. Using abs() here would silently invert
    # their contribution and misrepresent the variance structure (OECD §6.2).
    # Instead: exclude pillars with negative loadings and warn.
    neg_mask = loadings < 0
    if neg_mask.any():
        excluded = [p for p, neg in zip(pillars, neg_mask) if neg]
        logger.warning(
            "PCA: pillars with negative PC1 loadings excluded (inverse loading on "
            "stability dimension): %s", excluded
        )
        loadings = np.where(neg_mask, 0.0, loadings)
    if loadings.sum() < SMALL:
        logger.warning("PCA: all loadings excluded — falling back to equal weights")
        weights = np.array([1.0 / len(pillars)] * len(pillars))
    else:
        weights = loadings / loadings.sum()

    logger.info("  PCA weights: %s",
                {p: round(w, 4) for p, w in zip(pillars, weights)})

    # Weighted mean for all countries (including those with some NaN pillars)
    scores = {}
    for iso3, row in sub.iterrows():
        vals = row.values.astype(float)
        valid = ~np.isnan(vals)
        if valid.sum() == 0:
            scores[iso3] = float("nan")
        else:
            w_valid = weights[valid]
            w_valid = w_valid / w_valid.sum()
            scores[iso3] = float(np.dot(w_valid, vals[valid]))

    raw = pd.Series(scores)
    # Rescale to [0, 100]
    mn, mx = raw.min(skipna=True), raw.max(skipna=True)
    if mx == mn:
        return raw.where(raw.isna(), 50.0)
    return (raw - mn) / (mx - mn) * 100.0


def score_bod(pillar_wide: pd.DataFrame) -> pd.Series:
    """
    Benefit of Doubt composite (Cherchye et al. 2007).
    Each country receives the weight vector that maximises its own score,
    subject to global bounds [WEIGHT_MIN, WEIGHT_MAX] per pillar.
    Requires: pip install pulp
    """
    try:
        import pulp
    except ImportError:
        logger.error("BoD scoring requires 'pulp'. Install: pip install pulp")
        return pd.Series(float("nan"), index=pillar_wide.index)

    pillars = [p for p in PILLAR_DEFS if p in pillar_wide.columns]
    n_pillars = len(pillars)
    sub = pillar_wide[pillars] / 100.0  # normalise to [0, 1] for LP

    scores = {}
    for iso3, row in sub.iterrows():
        vals = row.values.astype(float)
        valid_own = [i for i in range(n_pillars) if not np.isnan(vals[i])]

        if not valid_own:
            scores[iso3] = float("nan")
            continue

        n_valid = len(valid_own)
        # Feasibility check: n_valid * WEIGHT_MAX >= 1 (can reach sum=1 at max weights).
        # With WEIGHT_MAX=0.25, n_valid < 4 makes the LP infeasible; fall back to
        # equal-weight mean. This also avoids the original bug where missing pillars
        # were included in the sum-to-1 constraint, forcing WEIGHT_MIN to be "paid"
        # for each missing pillar and artificially capping achievable scores.
        if n_valid * WEIGHT_MAX < 1.0 - 1e-9:
            scores[iso3] = float(np.mean([vals[i] for i in valid_own])) * 100.0
            continue

        prob = pulp.LpProblem(f"BoD_{iso3}", pulp.LpMaximize)
        # Weight variables only over valid pillars. Restricting the sum-to-1
        # constraint to valid pillars ensures the full budget is allocated to
        # pillars where the country has actual data.
        w = {i: pulp.LpVariable(f"w{i}", lowBound=WEIGHT_MIN, upBound=WEIGHT_MAX)
             for i in valid_own}

        prob += pulp.lpSum(w[i] for i in valid_own) == 1.0

        own_score = pulp.lpSum(w[i] * float(vals[i]) for i in valid_own)
        prob += own_score

        # Constraint: this weight vector cannot give any other country > 1.
        # Apply only to pillars where both countries have data.
        for other_iso3, other_row in sub.iterrows():
            if other_iso3 == iso3:
                continue
            other_vals = other_row.values.astype(float)
            common_valid = [i for i in valid_own if not np.isnan(other_vals[i])]
            if not common_valid:
                continue
            prob += pulp.lpSum(w[i] * float(other_vals[i]) for i in common_valid) <= 1.0

        prob.solve(pulp.PULP_CBC_CMD(msg=0))

        if pulp.LpStatus[prob.status] == "Optimal":
            scores[iso3] = float(own_score.value()) * 100.0
        else:
            scores[iso3] = float("nan")

    return pd.Series(scores)


def score_entropy(pillar_wide: pd.DataFrame) -> pd.Series:
    """
    Shannon entropy weights: pillars with higher cross-country variance
    receive higher weight.
    e_j = 1 - H_j / ln(n)  where H_j = -sum_i p_ij * ln(p_ij)
    p_ij = score_ij / sum_i score_ij
    Final weights normalised to sum to 1.
    """
    pillars = [p for p in PILLAR_DEFS if p in pillar_wide.columns]
    sub = pillar_wide[pillars].copy()
    n = len(sub)

    # Shift scores to strictly positive (entropy requires > 0)
    for col in sub.columns:
        mn = sub[col].min(skipna=True)
        if mn <= 0:
            sub[col] = sub[col] - mn + SMALL

    H = {}
    for col in sub.columns:
        col_vals = sub[col].fillna(sub[col].mean())   # fill NaN with column mean for entropy calc
        total = col_vals.sum()
        if total == 0:
            H[col] = 0.0
            continue
        p = col_vals / total
        p = p.clip(lower=SMALL)
        H[col] = float(-np.sum(p * np.log(p)))

    ln_n = np.log(n)
    e = {col: 1.0 - H[col] / ln_n for col in pillars}
    total_e = sum(e.values())
    if total_e == 0:
        weights = {col: 1.0 / len(pillars) for col in pillars}
    else:
        weights = {col: v / total_e for col, v in e.items()}

    logger.info("  Entropy weights: %s", {k: round(v, 4) for k, v in weights.items()})

    scores = {}
    for iso3, row in pillar_wide[pillars].iterrows():
        vals = row.values.astype(float)
        valid = ~np.isnan(vals)
        if valid.sum() == 0:
            scores[iso3] = float("nan")
        else:
            w_arr = np.array([weights[p] for p in pillars])
            w_valid = w_arr[valid] / w_arr[valid].sum()
            scores[iso3] = float(np.dot(w_valid, vals[valid]))

    return pd.Series(scores)


def score_geometric(pillar_wide: pd.DataFrame) -> pd.Series:
    """
    Geometric mean aggregation: exp(sum(w_i * ln(s_i))).
    Uses equal weights (1/7). Partial compensability — a near-zero pillar
    score dramatically drags down the composite regardless of other pillars.
    Scores are floored at SMALL before log to avoid log(0) = -inf.
    """
    pillars = [p for p in PILLAR_DEFS if p in pillar_wide.columns]
    n_pillars = len(pillars)
    w = np.array([1.0 / n_pillars] * n_pillars)
    sub = pillar_wide[pillars].copy()

    scores = {}
    for iso3, row in sub.iterrows():
        vals = row.values.astype(float)
        valid = ~np.isnan(vals)
        if valid.sum() == 0:
            scores[iso3] = float("nan")
            continue
        # Floor to SMALL: a score of 0 would give -inf in log
        vals_clipped = np.maximum(vals[valid], SMALL)
        w_valid = w[valid] / w[valid].sum()
        scores[iso3] = float(np.exp(np.dot(w_valid, np.log(vals_clipped))))

    return pd.Series(scores)


# ── Step 3: Rankings ───────────────────────────────────────────────────────────

def add_ranks(df: pd.DataFrame, score_cols: list[str]) -> pd.DataFrame:
    """Add rank columns for each score method (1 = best)."""
    for col in score_cols:
        df[f"{col}_rank"] = df[col].rank(ascending=False, method="min", na_option="bottom").astype("Int64")
    return df


# ── Excel formatting helpers ───────────────────────────────────────────────────

HEADER_FILL  = PatternFill(start_color="1B3A6B", end_color="1B3A6B", fill_type="solid")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
SUBHEAD_FILL = PatternFill(start_color="2E6DB4", end_color="2E6DB4", fill_type="solid")
SUBHEAD_FONT = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
CENTER       = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT         = Alignment(horizontal="left",   vertical="center")

# Green → Yellow → Red color scale (high score = green)
def _score_color_scale():
    return ColorScaleRule(
        start_type="min",        start_color="F26464",
        mid_type="percentile",   mid_value=50, mid_color="FFE664",
        end_type="max",          end_color="64C882",
    )

# Red → Yellow → Green (low rank number = better, so invert)
def _rank_color_scale():
    return ColorScaleRule(
        start_type="min",        start_color="64C882",
        mid_type="percentile",   mid_value=50, mid_color="FFE664",
        end_type="max",          end_color="F26464",
    )


def _style_header_row(ws, row: int, n_cols: int, fill=None, font=None):
    f = fill or HEADER_FILL
    fo = font or HEADER_FONT
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = f
        cell.font = fo
        cell.alignment = CENTER


def _set_col_width(ws, col: int, width: float):
    ws.column_dimensions[get_column_letter(col)].width = width


def _apply_cf(ws, col: int, min_row: int, max_row: int, rule):
    col_letter = get_column_letter(col)
    ws.conditional_formatting.add(f"{col_letter}{min_row}:{col_letter}{max_row}", rule)


def _fmt_score(cell):
    cell.number_format = "0.0"
    cell.alignment = CENTER


def _fmt_rank(cell):
    cell.number_format = "0"
    cell.alignment = CENTER


# ── Sheet builders ─────────────────────────────────────────────────────────────

def _write_overall_sheet(writer, overall: pd.DataFrame, score_cols: list[str],
                          name_map: dict, region_map: dict):
    """
    Sheet 1 — Overall Rankings
    Columns: Equal Rank | ISO3 | Country | Region |
             Equal Score | PCA Score | BoD Score | Entropy Score | Geometric Score |
             Equal Rank | PCA Rank | BoD Rank | Entropy Rank | Geometric Rank
    Sorted by Equal score descending. Auto-filter. Freeze cols A-D.
    """
    METHOD_LABELS = {
        "equal":     "Equal Weights",
        "pca":       "PCA Weights",
        "bod":       "Benefit of Doubt",
        "entropy":   "Entropy Weights",
        "geometric": "Geometric Mean",
    }

    df = overall.copy()
    df.insert(0, "Country",   df["iso3"].map(name_map).fillna(df["iso3"]))
    df.insert(0, "Region",    df["iso3"].map(region_map).fillna(""))
    df = df.sort_values("equal", ascending=False).reset_index(drop=True)
    df.insert(0, "Equal Rank", range(1, len(df) + 1))

    ws = writer.book.create_sheet("Overall Rankings")

    # ── Header row ─────────────────────────────────────────────────────────────
    headers = ["Rank", "ISO3", "Country", "Region"]
    score_header_cols = []
    rank_header_cols  = []

    for m in score_cols:
        headers.append(f"{METHOD_LABELS[m]}\nScore")
        score_header_cols.append(len(headers))
    for m in score_cols:
        headers.append(f"{METHOD_LABELS[m]}\nRank")
        rank_header_cols.append(len(headers))

    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER

    ws.row_dimensions[1].height = 32

    # ── Data rows ──────────────────────────────────────────────────────────────
    data_cols = (
        ["Equal Rank", "iso3", "Country", "Region"]
        + score_cols
        + [f"{m}_rank" for m in score_cols]
    )

    for r_idx, (_, row) in enumerate(df[data_cols].iterrows(), 2):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val if pd.notna(val) else None)
            if c_idx in score_header_cols:
                _fmt_score(cell)
            elif c_idx in rank_header_cols or c_idx == 1:
                _fmt_rank(cell)
            else:
                cell.alignment = LEFT

    n_data = len(df)

    # ── Conditional formatting ──────────────────────────────────────────────────
    for c in score_header_cols:
        _apply_cf(ws, c, 2, n_data + 1, _score_color_scale())
    for c in rank_header_cols:
        _apply_cf(ws, c, 2, n_data + 1, _rank_color_scale())

    # ── Column widths ──────────────────────────────────────────────────────────
    widths = [6, 7, 22, 10] + [14] * len(score_cols) + [10] * len(score_cols)
    for c, w in enumerate(widths, 1):
        _set_col_width(ws, c, w)

    # ── Freeze panes + auto-filter ─────────────────────────────────────────────
    ws.freeze_panes = "E2"    # freeze top row + first 4 columns
    ws.auto_filter.ref = ws.dimensions

    # ── Instruction note in cell A1 tooltip area ────────────────────────────────
    ws["A1"].comment = None   # clear any existing
    # (openpyxl comments require an extra dep; skip — filter arrows are self-evident)


def _write_pillar_sheet(writer, pillar_wide: pd.DataFrame,
                         name_map: dict, region_map: dict):
    """
    Sheet 2 — Pillar Scores
    One row per country. Columns: ISO3 | Country | Region | <Pillar Name> Score | <Pillar Name> Rank | ...
    Pillar letters expanded to full names. Sorted by equal-weight composite descending.
    Auto-filter. Freeze cols A-C.
    """
    df = pillar_wide.reset_index().copy()
    df.insert(1, "Country", df["iso3"].map(name_map).fillna(df["iso3"]))
    df.insert(2, "Region",  df["iso3"].map(region_map).fillna(""))

    # Add equal-weight composite for default sort
    pillar_ids = [p for p in PILLAR_DEFS if p in pillar_wide.columns]
    df["_composite"] = df[pillar_ids].mean(axis=1, skipna=True)
    df = df.sort_values("_composite", ascending=False).reset_index(drop=True)
    df = df.drop(columns=["_composite"])

    # Compute ranks
    for p in pillar_ids:
        df[f"{p}_rank"] = df[p].rank(ascending=False, method="min", na_option="bottom").astype("Int64")

    ws = writer.book.create_sheet("Pillar Scores")

    # Build interleaved columns: Score | Rank for each pillar
    fixed_cols    = ["iso3", "Country", "Region"]
    score_ci      = []   # 1-based column indices of score columns
    rank_ci       = []

    headers = ["ISO3", "Country", "Region"]
    data_col_keys = ["iso3", "Country", "Region"]

    for p in pillar_ids:
        full_name = PILLAR_DEFS[p]
        headers.append(f"{full_name}\nScore")
        score_ci.append(len(headers))
        data_col_keys.append(p)

        headers.append(f"{full_name}\nRank")
        rank_ci.append(len(headers))
        data_col_keys.append(f"{p}_rank")

    # Header
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
    ws.row_dimensions[1].height = 32

    # Data
    for r_idx, (_, row) in enumerate(df[data_col_keys].iterrows(), 2):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val if pd.notna(val) else None)
            if c_idx in score_ci:
                _fmt_score(cell)
            elif c_idx in rank_ci:
                _fmt_rank(cell)
            else:
                cell.alignment = LEFT

    n_data = len(df)
    for c in score_ci:
        _apply_cf(ws, c, 2, n_data + 1, _score_color_scale())
    for c in rank_ci:
        _apply_cf(ws, c, 2, n_data + 1, _rank_color_scale())

    # Column widths: ISO3=7, Country=22, Region=10, then alternating 14/8 per pillar
    widths = [7, 22, 10] + [14, 8] * len(pillar_ids)
    for c, w in enumerate(widths, 1):
        _set_col_width(ws, c, w)

    ws.freeze_panes = "D2"
    ws.auto_filter.ref = ws.dimensions


def _write_indicator_sheet(writer, norm_df: pd.DataFrame,
                            pillar_map: dict, name_map: dict):
    """
    Sheet 3 — Indicator Scores
    Long format: ISO3 | Country | Pillar | Indicator | Score (0-100)
    Grouped by pillar (full name), sorted alphabetically within pillar.
    Auto-filter. Freeze cols A-B.
    """
    # Build indicator → pillar lookup (use first pillar if multi-pillar)
    ind_to_pillar = {}
    for p, vars_in in pillar_map.items():
        for v in vars_in:
            if v not in ind_to_pillar:
                ind_to_pillar[v] = p

    df = norm_df.copy()
    df["Country"]     = df["iso3"].map(name_map).fillna(df["iso3"])
    df["Pillar"]      = df["variable_name"].map(
        lambda v: PILLAR_DEFS.get(ind_to_pillar.get(v, ""), ind_to_pillar.get(v, ""))
    )
    df["Pillar Code"] = df["variable_name"].map(lambda v: ind_to_pillar.get(v, ""))
    df = df.sort_values(["Pillar Code", "variable_name", "iso3"]).reset_index(drop=True)

    out = df[["iso3", "Country", "Pillar", "variable_name", "norm_score"]].copy()
    out.columns = ["ISO3", "Country", "Pillar", "Indicator", "Score (0-100)"]

    ws = writer.book.create_sheet("Indicator Scores")

    headers = list(out.columns)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER

    score_col = len(headers)  # last column

    for r_idx, (_, row) in enumerate(out.iterrows(), 2):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val if pd.notna(val) else None)
            if c_idx == score_col:
                _fmt_score(cell)
            else:
                cell.alignment = LEFT

    n_data = len(out)
    _apply_cf(ws, score_col, 2, n_data + 1, _score_color_scale())

    widths = [7, 22, 26, 34, 14]
    for c, w in enumerate(widths, 1):
        _set_col_width(ws, c, w)

    ws.freeze_panes = "C2"
    ws.auto_filter.ref = ws.dimensions


def _write_weights_sheet(writer, method_weights: dict):
    """
    Sheet 4 — Method Weights
    Shows what weight each method assigned to each pillar.
    Rows = pillars (full names). Columns = methods.
    """
    METHOD_LABELS = {
        "equal":     "Equal Weights",
        "pca":       "PCA Weights",
        "bod_min":   "BoD Min Bound",
        "bod_max":   "BoD Max Bound",
        "entropy":   "Entropy Weights",
        "geometric": "Geometric Mean",
    }

    ws = writer.book.create_sheet("Method Weights")

    methods = [m for m in METHOD_LABELS if m in method_weights]
    headers = ["Pillar", "Pillar Code"] + [METHOD_LABELS[m] for m in methods]

    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER

    for r_idx, (pillar_id, pillar_name) in enumerate(PILLAR_DEFS.items(), 2):
        ws.cell(row=r_idx, column=1, value=pillar_name).alignment = LEFT
        ws.cell(row=r_idx, column=2, value=pillar_id).alignment = CENTER
        for c_idx, method in enumerate(methods, 3):
            w = method_weights.get(method, {}).get(pillar_id, None)
            cell = ws.cell(row=r_idx, column=c_idx, value=round(w, 4) if w is not None else None)
            cell.number_format = "0.0000"
            cell.alignment = CENTER

    # Style the fixed-weight row note
    note_row = len(PILLAR_DEFS) + 3
    ws.cell(row=note_row, column=1,
            value="Note: BoD weights are per-country optimal — min/max bounds shown here.").font = Font(italic=True, size=9)

    widths = [28, 10] + [16] * len(methods)
    for c, w in enumerate(widths, 1):
        _set_col_width(ws, c, w)

    ws.freeze_panes = "C2"


def _compute_method_weights(pillar_wide: pd.DataFrame, pillar_map: dict) -> dict:
    """Re-derive weights for the weights sheet (equal, PCA, entropy, BoD bounds, geometric)."""
    from sklearn.decomposition import PCA
    from sklearn.preprocessing import StandardScaler

    pillars = [p for p in PILLAR_DEFS if p in pillar_wide.columns]
    weights = {}

    # Equal (active preset)
    preset = WEIGHT_PRESETS[ACTIVE_PRESET]
    weights["equal"] = {p: preset.get(p, 1/len(pillars)) for p in pillars}

    # PCA
    sub = pillar_wide[pillars]
    complete = sub.dropna()
    if len(complete) >= len(pillars):
        X = StandardScaler().fit_transform(complete.values)
        pca = PCA(n_components=1).fit(X)
        loadings = pca.components_[0]
        if (loadings < 0).sum() > len(loadings) / 2:
            loadings = -loadings
        neg_mask = loadings < 0
        loadings_clean = np.where(neg_mask, 0.0, loadings)
        w_pca = loadings_clean / loadings_clean.sum() if loadings_clean.sum() > SMALL else np.ones(len(pillars)) / len(pillars)
        weights["pca"] = {p: float(w) for p, w in zip(pillars, w_pca)}

    # Entropy
    sub2 = pillar_wide[pillars].copy()
    for col in sub2.columns:
        mn = sub2[col].min(skipna=True)
        if mn <= 0:
            sub2[col] = sub2[col] - mn + SMALL
    n = len(sub2)
    H = {}
    for col in pillars:
        col_vals = sub2[col].fillna(sub2[col].mean())
        total = col_vals.sum()
        p = (col_vals / total).clip(lower=SMALL)
        H[col] = float(-np.sum(p * np.log(p)))
    ln_n = np.log(n)
    e = {col: 1.0 - H[col] / ln_n for col in pillars}
    total_e = sum(e.values())
    weights["entropy"] = {col: v / total_e for col, v in e.items()} if total_e > 0 else {p: 1/len(pillars) for p in pillars}

    # BoD bounds
    weights["bod_min"] = {p: WEIGHT_MIN for p in pillars}
    weights["bod_max"] = {p: WEIGHT_MAX for p in pillars}

    # Geometric (equal weights in log space)
    weights["geometric"] = {p: 1/len(pillars) for p in pillars}

    return weights


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    global ACTIVE_PRESET   # must be declared before any reference to it in this scope
    import argparse
    parser = argparse.ArgumentParser(description="AII scoring stage")
    parser.add_argument(
        "--preset", default=None,
        help=f"Weight preset to use. Available: {list(WEIGHT_PRESETS.keys())}. "
             f"Overrides ACTIVE_PRESET in constants.py (currently '{ACTIVE_PRESET}')."
    )
    args = parser.parse_args()
    if args.preset is not None:
        if args.preset not in WEIGHT_PRESETS:
            logger.error("Unknown preset '%s'. Available: %s", args.preset, list(WEIGHT_PRESETS.keys()))
            sys.exit(1)
        ACTIVE_PRESET = args.preset

    logger.info("Active weight preset: %s", ACTIVE_PRESET)

    t0 = _time.time()
    logger.info("Step 1: Load normalized data from %s", INPUT_FILE)

    if not INPUT_FILE.exists():
        logger.error("Input file not found: %s — run 03_normalize.py first.", INPUT_FILE)
        sys.exit(1)

    norm_df = pd.read_excel(INPUT_FILE, sheet_name="norm_data")
    logger.info("  Loaded %d rows, %d countries, %d indicators",
                len(norm_df), norm_df["iso3"].nunique(), norm_df["variable_name"].nunique())

    # Pivot to wide
    norm_wide = norm_df.pivot(index="iso3", columns="variable_name", values="norm_score")
    norm_wide.columns.name = None

    logger.info("Step 2: Load pillar indicator assignments")
    pillar_map = load_pillar_map()
    for p, vars_in in pillar_map.items():
        logger.info("  Pillar %s: %d indicators", p, len(vars_in))

    logger.info("Step 3: Compute pillar scores (arithmetic mean within pillar)")
    pillar_wide = compute_pillar_scores(norm_wide, pillar_map)

    pillar_long = pillar_wide.reset_index().melt(
        id_vars="iso3", var_name="pillar", value_name="pillar_score"
    )

    logger.info("Step 4: Compute overall scores — 5 methods")
    methods_map = {
        "equal":    score_equal,
        "pca":      score_pca,
        "bod":      score_bod,
        "entropy":  score_entropy,
        "geometric": score_geometric,
    }

    overall = pillar_wide.copy()
    for method, fn in methods_map.items():
        logger.info("  Method: %s", method)
        try:
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                overall[method] = fn(pillar_wide)
        except Exception as exc:
            logger.error("  %s failed: %s", method, exc)
            overall[method] = float("nan")

    score_cols = list(methods_map.keys())
    overall = overall.reset_index()
    overall = add_ranks(overall, score_cols)

    # Summary stats
    for method in score_cols:
        col_scores = overall[method].dropna()
        if len(col_scores) == 0:
            logger.info("  %-12s  no valid scores", method)
        else:
            logger.info(
                "  %-12s  mean=%.1f  min=%.1f (%s)  max=%.1f (%s)",
                method,
                col_scores.mean(),
                col_scores.min(),
                overall.loc[col_scores.idxmin(), "iso3"],
                col_scores.max(),
                overall.loc[col_scores.idxmax(), "iso3"],
            )

    logger.info("Step 5: Write formatted output to %s", OUTPUT_FILE)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # Load country names and regions for display
    from models.countries import COUNTRIES
    name_map   = {k: v["name"]   for k, v in COUNTRIES.items()}
    region_map = {k: v["region"] for k, v in COUNTRIES.items()}

    # Capture method weights for the weights sheet
    method_weights = _compute_method_weights(pillar_wide, pillar_map)

    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        _write_overall_sheet(writer, overall, score_cols, name_map, region_map)
        _write_pillar_sheet(writer, pillar_wide, name_map, region_map)
        _write_indicator_sheet(writer, norm_df, pillar_map, name_map)
        _write_weights_sheet(writer, method_weights)

    # Machine-readable CSV for downstream consumption (05_robustness, 06_qualitative).
    # Avoids parsing the formatted Excel which has multi-row headers and merged cells.
    csv_file = OUTPUT_DIR / "04_scores_raw.csv"
    csv_df = overall[["iso3"] + score_cols + [f"{m}_rank" for m in score_cols]].copy()
    pillar_for_csv = pillar_wide.reset_index()[["iso3"] + [p for p in PILLAR_DEFS if p in pillar_wide.columns]]
    csv_df = csv_df.merge(pillar_for_csv, on="iso3", how="left")
    csv_df.to_csv(csv_file, index=False)
    logger.info("  CSV: %s (%d rows)", csv_file, len(csv_df))

    logger.info("Done in %.1fs", _time.time() - t0)
    logger.info("  %d countries scored, 5 methods applied", len(overall))
    logger.info("  Sheets: Overall Rankings | Pillar Scores | Indicator Scores | Method Weights")


if __name__ == "__main__":
    main()
