"""
06_qualitative.py — Assemble the full results bundle with qualitative content.

Input:  data/04_scores.xlsx
        data/02_clean.xlsx  (fill_log, clean_data)
        data/05_robustness.json
        indicators_list/*.yaml
        config.py  (PillarRegistry)
Output: data/06_results.json   — master results bundle (consumed by dashboard + AI)
        data/06_qualitative.xlsx — human-readable qualitative reference

The results bundle structure:
  {
    "run":     { project, date, n_entities, scoring_methods, ... }
    "pillars": { A: { name, justification, indicators: [...] }, ... }
    "indicators": { var_name: { display_name, justification, pillar, polarity, ... }, ... }
    "countries": [
      {
        "iso3": "NGA", "name": "Nigeria", "region": "West",
        "island_state": false,
        "data_quality": { "pct_real": 0.82, "n_filled": 6, "fill_flag": "Good" },
        "scores": { "equal": 44.2, "pca": 41.1, ... },
        "ranks":  { "equal": 22, ... },
        "pillar_scores": { "A": 38.1, "B": 51.2, ... },
        "pillar_ranks":  { "A": 30, ... },
        "indicators": { "va_estimate": { "score": 42.1, "filled": false }, ... },
        "peer_rank": { "region": "West", "rank_in_region": 4, "n_in_region": 15 },
        "confidence_band": { "equal_low": 40.1, "equal_high": 48.3 }
      }, ...
    ]
  }

Run from the project root:
    python 06_qualitative.py
"""

import sys
import logging
import json
import time as _time
from datetime import date

import numpy as np
import pandas as pd
import yaml
from pathlib import Path
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from constants import PILLAR_DEFS

# ── Logging ────────────────────────────────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format="%(levelname)-8s %(message)s",
    stream=sys.stdout,
)
logger = logging.getLogger(__name__)

# ── Paths ──────────────────────────────────────────────────────────────────────

SCORES_FILE    = Path("data/04_scores.xlsx")
CLEAN_FILE     = Path("data/02_clean.xlsx")
ROBUSTNESS_JSON = Path("data/05_robustness.json")
INDICATORS_DIR = Path("indicators_list")
OUTPUT_DIR     = Path("data")
OUTPUT_JSON    = OUTPUT_DIR / "06_results.json"
OUTPUT_XLSX    = OUTPUT_DIR / "06_qualitative.xlsx"

# ── Constants ──────────────────────────────────────────────────────────────────

METHODS     = ["equal", "pca", "bod", "entropy", "geometric"]
ISLAND_SET  = {"SYC", "MUS", "COM", "CPV", "STP", "MDG"}

METHOD_DESCRIPTIONS = {
    "equal": {
        "label": "Equal Weights",
        "short": "Each pillar weighted equally (1/7). Simple, transparent, fully compensatory.",
        "interpretation": (
            "A score of 100 means the country performs best in the sample on the active weight preset. "
            "A score of 0 means it performs worst. Scores are relative to the 54-country sample — "
            "not an absolute measure of stability. Because weights are equal, a very poor performance "
            "in one pillar can be offset by strong performance in others."
        ),
        "strength": "Most transparent and easily explained.",
        "limitation": "Fully compensatory — a war-torn country with strong GDP still scores well.",
    },
    "pca": {
        "label": "PCA Weights",
        "short": "First principal component loadings as pillar weights. Data-driven.",
        "interpretation": (
            "Weights are derived from where countries actually differ most. "
            "Pillars that vary most across countries receive more weight. "
            "Scores are rescaled to [0, 100] after PCA. The absolute score value has less meaning "
            "than the ranking — two countries close in score may not be meaningfully different."
        ),
        "strength": "Reflects the actual variance structure of the data.",
        "limitation": "Weights change if countries are added or removed. Not comparable across versions.",
    },
    "bod": {
        "label": "Benefit of Doubt",
        "short": "Each country receives the weight vector that maximises its own score (within bounds).",
        "interpretation": (
            "BoD scores are best interpreted as 'ceiling scores' — the best a country can plausibly "
            "claim given any reasonable weighting. A country ranked low under BoD has few genuine "
            "strengths even when measured most charitably. Scores tend to cluster higher than equal-weight "
            "because each country is given its best case."
        ),
        "strength": "Eliminates debate about weight choice for countries doing best on different dimensions.",
        "limitation": "BoD scores are not directly comparable across countries — each uses different weights.",
    },
    "entropy": {
        "label": "Entropy Weights",
        "short": "Pillars with more cross-country variance receive higher weight.",
        "interpretation": (
            "Entropy weights reward informative pillars — those where countries genuinely differ. "
            "A pillar where all countries score similarly is down-weighted. "
            "In practice, Health and Governance tend to receive higher entropy weights for Africa "
            "because variation is high. Economic indicators tend to be lower due to fill effects."
        ),
        "strength": "Objective, data-driven. No expert judgment required.",
        "limitation": "Sensitive to sample composition. Adding countries changes weights.",
    },
    "geometric": {
        "label": "Geometric Mean",
        "short": "exp(Σ wᵢ·ln(sᵢ)). Partially non-compensatory.",
        "interpretation": (
            "The geometric mean penalises uneven development. A country that scores very low on "
            "any single pillar — even if it excels elsewhere — will have a significantly lower "
            "geometric score than equal-weight score. This is methodologically appropriate for "
            "stability: a country at war is not 'stable' just because its economy is growing. "
            "The gap between a country's equal-weight and geometric score reveals the cost of "
            "its weakest pillar."
        ),
        "strength": "Most defensible for stability indices. Preferred by Munda & Nardo (2009).",
        "limitation": "Harsh on countries with one very weak pillar, even if improving rapidly.",
    },
}

# ── Data loaders ───────────────────────────────────────────────────────────────

def load_indicator_meta() -> dict:
    """Returns {variable_name: full metadata dict} from indicator YAMLs."""
    meta = {}
    for path in sorted(INDICATORS_DIR.glob("*.yaml")):
        with open(path, "r", encoding="utf-8") as f:
            data = yaml.safe_load(f)
        if not data:
            continue
        for ind in data:
            meta[ind["variable_name"]] = ind
    return meta


def load_pillar_justifications() -> dict:
    """Load pillar justifications from context/pillar_justifications.yaml.

    This local file replaces the prior dependency on ../Athena/presets/aii.yaml,
    which is an external path that does not exist in the deployed environment.
    context/pillar_justifications.yaml is a flat {pillar_id: justification_text}
    YAML owned and maintained within this repository.
    """
    local_path = Path("context/pillar_justifications.yaml")
    fallback = {k: "" for k in PILLAR_DEFS}
    if not local_path.exists():
        logger.warning("  context/pillar_justifications.yaml not found — pillar justifications empty")
        return fallback
    with open(local_path, "r", encoding="utf-8") as f:
        d = yaml.safe_load(f) or {}
    return {pillar_id: str(d.get(pillar_id, "")).strip() for pillar_id in PILLAR_DEFS}


def load_scores_data():
    """Load all score sheets from 04_scores.xlsx."""
    overall   = pd.read_excel(SCORES_FILE, sheet_name="Overall Rankings")
    pillar_w  = pd.read_excel(SCORES_FILE, sheet_name="Pillar Scores")
    ind_scores = pd.read_excel(SCORES_FILE, sheet_name="Indicator Scores")

    # Normalise headers
    overall.columns   = [c.replace("\n", " ").strip() for c in overall.columns]
    pillar_w.columns  = [c.replace("\n", " ").strip() for c in pillar_w.columns]

    col_map = {
        "Equal Weights Score":    "equal",
        "PCA Weights Score":      "pca",
        "Benefit of Doubt Score": "bod",
        "Entropy Weights Score":  "entropy",
        "Geometric Mean Score":   "geometric",
        "Equal Weights Rank":     "equal_rank",
        "PCA Weights Rank":       "pca_rank",
        "Benefit of Doubt Rank":  "bod_rank",
        "Entropy Weights Rank":   "entropy_rank",
        "Geometric Mean Rank":    "geometric_rank",
        "ISO3": "iso3",
    }
    overall  = overall.rename(columns=col_map)
    pillar_w = pillar_w.rename(columns={"ISO3": "iso3"})

    return overall, pillar_w, ind_scores


# ── Data quality per country ───────────────────────────────────────────────────

def compute_data_quality(fill_log: pd.DataFrame, indicator_meta: dict,
                          n_indicators: int) -> dict:
    """
    Returns {iso3: {pct_real, n_filled, fill_flag, filled_indicators}}.
    fill_flag: "Excellent" (0%) | "Good" (<15%) | "Moderate" (<30%) | "Caution" (>=30%)
    """
    filled_by_country: dict[str, list] = {}
    for _, row in fill_log.iterrows():
        iso3 = row.get("iso3")
        var  = row.get("variable_name")
        if iso3 and var:
            filled_by_country.setdefault(iso3, []).append(var)

    result = {}
    for iso3, filled in filled_by_country.items():
        n_filled = len(set(filled))
        pct_real = round(1 - n_filled / n_indicators, 4)
        flag = (
            "Excellent" if n_filled == 0 else
            "Good"      if n_filled / n_indicators < 0.15 else
            "Moderate"  if n_filled / n_indicators < 0.30 else
            "Caution"
        )
        result[iso3] = {
            "pct_real":           pct_real,
            "n_filled":           n_filled,
            "n_total_indicators": n_indicators,
            "fill_flag":          flag,
            "filled_indicators":  sorted(set(filled)),
        }
    return result


def compute_confidence_band(overall_score: float, dq: dict,
                             method: str = "equal") -> dict:
    """
    Simple confidence band based on fill rate.
    Width scales with % filled: 0% filled = ±2 pts, 30%+ filled = ±10 pts.
    """
    if overall_score is None or np.isnan(overall_score):
        return {"low": None, "high": None, "width": None}
    pct_filled = 1 - dq.get("pct_real", 1.0)
    half_width = 2.0 + pct_filled * 26.7   # linear: 0%→2, 30%→10
    low  = max(0.0, round(overall_score - half_width, 1))
    high = min(100.0, round(overall_score + half_width, 1))
    return {"low": low, "high": high, "width": round(half_width * 2, 1)}


def compute_peer_ranks(overall: pd.DataFrame, region_map: dict) -> dict:
    """
    Returns {iso3: {region, rank_in_region, n_in_region}}.
    """
    result = {}
    overall_w_region = overall.copy()
    overall_w_region["_region"] = overall_w_region["iso3"].map(region_map)

    for region, grp in overall_w_region.groupby("_region"):
        grp_sorted = grp.sort_values("equal", ascending=False).reset_index(drop=True)
        for rank_idx, (_, row) in enumerate(grp_sorted.iterrows(), 1):
            result[row["iso3"]] = {
                "region":         region,
                "rank_in_region": rank_idx,
                "n_in_region":    len(grp_sorted),
            }
    return result


# ── Build results bundle ───────────────────────────────────────────────────────

def build_results_bundle(
    overall: pd.DataFrame,
    pillar_wide: pd.DataFrame,
    ind_scores_df: pd.DataFrame,
    indicator_meta: dict,
    pillar_just: dict,
    dq_map: dict,
    peer_ranks: dict,
    name_map: dict,
    region_map: dict,
) -> dict:

    pillar_wide = pillar_wide.set_index("iso3") if "iso3" in pillar_wide.columns else pillar_wide

    # Build pillar indicator lookup
    pillar_indicators: dict[str, list] = {k: [] for k in PILLAR_DEFS}
    for var, meta in indicator_meta.items():
        for p in meta.get("pillars", []):
            if p in pillar_indicators:
                pillar_indicators[p].append(var)

    bundle: dict = {
        "run": {
            "project":         "African Stability Index",
            "version":         "1.0",
            "date":            str(date.today()),
            "n_entities":      len(overall),
            "scoring_methods": METHODS,
            "scope":           "54 African Union member states",
            "note": (
                "Scores are relative to the 54-country sample. "
                "Higher score = more stable. All scores in [0, 100]."
            ),
        },
        "method_guide": METHOD_DESCRIPTIONS,
        "pillars": {
            pillar_id: {
                "name":         PILLAR_DEFS[pillar_id],
                "justification": pillar_just.get(pillar_id, ""),
                "indicators":   pillar_indicators.get(pillar_id, []),
                "n_indicators": len(pillar_indicators.get(pillar_id, [])),
            }
            for pillar_id in PILLAR_DEFS
        },
        "indicators": {
            var: {
                "display_name":  meta.get("display_name", var),
                "justification": str(meta.get("justification", "")).strip(),
                "series_code":   meta.get("series_code", ""),
                "source":        meta.get("source", ""),
                "database":      meta.get("database", ""),
                "pillars":       meta.get("pillars", []),
                "polarity":      meta.get("polarity", "positive"),
                "log_transform": bool(meta.get("log_transform", False)),
                "aggregation":   meta.get("aggregation", ""),
                "year_start":    meta.get("year_start"),
                "year_end":      meta.get("year_end"),
                "role":          meta.get("role", "scoring"),
            }
            for var, meta in indicator_meta.items()
        },
        "countries": [],
    }

    # Pivot indicator scores to wide for fast lookup
    ind_wide = ind_scores_df.pivot_table(
        index="ISO3", columns="Indicator", values="Score (0-100)", aggfunc="first"
    ) if "ISO3" in ind_scores_df.columns else pd.DataFrame()

    for _, row in overall.iterrows():
        iso3 = row["iso3"]
        dq   = dq_map.get(iso3, {
            "pct_real": 1.0, "n_filled": 0,
            "n_total_indicators": len(indicator_meta),
            "fill_flag": "Excellent", "filled_indicators": [],
        })
        peer = peer_ranks.get(iso3, {})

        # Pillar scores + ranks
        pillar_scores = {}
        pillar_ranks  = {}
        for p in PILLAR_DEFS:
            full_name = PILLAR_DEFS[p]
            score_col = f"{full_name} Score"
            rank_col  = f"{full_name} Rank"
            if score_col in pillar_wide.columns:
                v = pillar_wide.loc[iso3, score_col] if iso3 in pillar_wide.index else None
                pillar_scores[p] = round(float(v), 2) if v is not None and pd.notna(v) else None
            if rank_col in pillar_wide.columns:
                v = pillar_wide.loc[iso3, rank_col] if iso3 in pillar_wide.index else None
                pillar_ranks[p] = int(v) if v is not None and pd.notna(v) else None

        # Indicator scores for this country
        ind_data = {}
        if iso3 in ind_wide.index:
            for var in indicator_meta:
                if var in ind_wide.columns:
                    val = ind_wide.loc[iso3, var]
                    ind_data[var] = {
                        "score":  round(float(val), 2) if pd.notna(val) else None,
                        "filled": var in dq.get("filled_indicators", []),
                    }

        eq_score = row.get("equal")
        country_entry = {
            "iso3":         iso3,
            "name":         name_map.get(iso3, iso3),
            "region":       region_map.get(iso3, ""),
            "island_state": iso3 in ISLAND_SET,
            "data_quality": dq,
            "scores": {
                m: round(float(row[m]), 2) if pd.notna(row.get(m)) else None
                for m in METHODS
            },
            "ranks": {
                m: int(row[f"{m}_rank"]) if pd.notna(row.get(f"{m}_rank")) else None
                for m in METHODS
            },
            "pillar_scores":    pillar_scores,
            "pillar_ranks":     pillar_ranks,
            "indicators":       ind_data,
            "peer_rank":        peer,
            "confidence_band":  compute_confidence_band(
                float(eq_score) if pd.notna(eq_score) else None, dq
            ),
        }
        bundle["countries"].append(country_entry)

    # Sort countries by equal-weight score descending
    bundle["countries"].sort(
        key=lambda c: c["scores"].get("equal") or 0, reverse=True
    )
    return bundle


# ── Excel output ───────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="1B3A6B", end_color="1B3A6B", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT        = Alignment(horizontal="left",   vertical="center")


def _write_country_profiles(wb, bundle: dict):
    """Sheet: Country Data Quality — iso3, name, fill stats, confidence band."""
    ws = wb.create_sheet("Country Profiles")
    headers = ["ISO3", "Country", "Region", "Island State",
               "% Real Data", "Indicators Filled", "Fill Flag",
               "Equal Score", "Equal Rank", "Confidence Low", "Confidence High",
               "Peer Rank in Region", "Region Size"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL; cell.font = HEADER_FONT; cell.alignment = CENTER
    ws.row_dimensions[1].height = 28

    flag_colors = {"Excellent": "eefaf4", "Good": "f0fff4", "Moderate": "fff8e1", "Caution": "fdecea"}
    for r, c in enumerate(bundle["countries"], 2):
        dq   = c["data_quality"]
        band = c["confidence_band"]
        peer = c["peer_rank"]
        row  = [
            c["iso3"], c["name"], c["region"],
            "Yes" if c["island_state"] else "",
            round(dq["pct_real"] * 100, 1),
            dq["n_filled"], dq["fill_flag"],
            c["scores"].get("equal"), c["ranks"].get("equal"),
            band.get("low"), band.get("high"),
            peer.get("rank_in_region"), peer.get("n_in_region"),
        ]
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=col, value=val if val is not None else "")
            if headers[col - 1] == "Fill Flag":
                cell.fill = PatternFill(
                    start_color=flag_colors.get(str(val), "FFFFFF"),
                    end_color=flag_colors.get(str(val), "FFFFFF"), fill_type="solid")
                cell.alignment = CENTER
            elif headers[col - 1] in ("% Real Data", "Equal Score",
                                       "Confidence Low", "Confidence High"):
                cell.number_format = "0.0"; cell.alignment = CENTER
            elif col >= 8:
                cell.alignment = CENTER
            else:
                cell.alignment = LEFT

    widths = [7, 22, 10, 10, 12, 16, 10, 12, 10, 14, 14, 18, 12]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "E2"
    ws.auto_filter.ref = ws.dimensions


def _write_indicator_reference(wb, bundle: dict):
    """Sheet: Indicator Reference — full metadata table."""
    ws = wb.create_sheet("Indicator Reference")
    headers = ["Variable Name", "Display Name", "Pillars", "Source", "Series Code",
               "Polarity", "Log Transform", "Aggregation", "Year Start", "Year End",
               "Justification"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL; cell.font = HEADER_FONT; cell.alignment = CENTER
    ws.row_dimensions[1].height = 28

    for r, (var, meta) in enumerate(bundle["indicators"].items(), 2):
        row = [
            var,
            meta["display_name"],
            ", ".join(meta["pillars"]),
            meta["source"],
            meta["series_code"],
            meta["polarity"],
            "Yes" if meta["log_transform"] else "No",
            meta["aggregation"],
            meta["year_start"],
            meta["year_end"],
            meta["justification"],
        ]
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = LEFT if c == len(headers) else CENTER
            if c == len(headers):
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    widths = [22, 40, 10, 10, 22, 10, 12, 18, 10, 10, 60]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def _write_pillar_reference(wb, bundle: dict):
    """Sheet: Pillar Reference — justifications and indicator membership."""
    ws = wb.create_sheet("Pillar Reference")
    headers = ["Pillar ID", "Pillar Name", "Indicators (count)", "Justification"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL; cell.font = HEADER_FONT; cell.alignment = CENTER
    ws.row_dimensions[1].height = 28

    for r, (pid, pdata) in enumerate(bundle["pillars"].items(), 2):
        row = [pid, pdata["name"], pdata["n_indicators"], pdata["justification"]]
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c == 4:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                ws.row_dimensions[r].height = 60
            else:
                cell.alignment = CENTER

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 80


def _write_method_guide(wb):
    """Sheet: Method Guide — plain-English interpretation of each method."""
    ws = wb.create_sheet("Method Guide")
    headers = ["Method", "Label", "One-line Summary", "How to Interpret",
               "Strength", "Limitation"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL; cell.font = HEADER_FONT; cell.alignment = CENTER
    ws.row_dimensions[1].height = 28

    for r, (method, desc) in enumerate(METHOD_DESCRIPTIONS.items(), 2):
        row = [method, desc["label"], desc["short"],
               desc["interpretation"], desc["strength"], desc["limitation"]]
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 80

    for c, w in enumerate([14, 20, 40, 80, 40, 50], 1):
        ws.column_dimensions[get_column_letter(c)].width = w


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    t0 = _time.time()

    for f in [SCORES_FILE, CLEAN_FILE]:
        if not f.exists():
            logger.error("Required file not found: %s", f)
            sys.exit(1)

    logger.info("Step 1: Load scores data")
    overall, pillar_wide, ind_scores_df = load_scores_data()
    logger.info("  %d countries loaded", len(overall))

    logger.info("Step 2: Load indicator metadata")
    indicator_meta = load_indicator_meta()
    logger.info("  %d indicators defined", len(indicator_meta))

    logger.info("Step 3: Load pillar justifications")
    pillar_just = load_pillar_justifications()

    logger.info("Step 4: Load fill log and compute data quality")
    fill_log = pd.read_excel(CLEAN_FILE, sheet_name="fill_log")
    from models.countries import COUNTRIES
    name_map   = {k: v["name"]   for k, v in COUNTRIES.items()}
    region_map = {k: v["region"] for k, v in COUNTRIES.items()}
    dq_map = compute_data_quality(fill_log, indicator_meta, len(indicator_meta))
    n_caution = sum(1 for v in dq_map.values() if v["fill_flag"] == "Caution")
    logger.info("  %d countries with 'Caution' fill rate (>30%% filled)", n_caution)

    logger.info("Step 5: Compute peer ranks")
    peer_ranks = compute_peer_ranks(overall, region_map)

    logger.info("Step 6: Build results bundle")
    bundle = build_results_bundle(
        overall, pillar_wide, ind_scores_df, indicator_meta,
        pillar_just, dq_map, peer_ranks, name_map, region_map,
    )

    logger.info("Step 7: Write JSON results bundle to %s", OUTPUT_JSON)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(bundle, f, indent=2, ensure_ascii=False)

    logger.info("Step 8: Write qualitative Excel to %s", OUTPUT_XLSX)
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)
    _write_country_profiles(wb, bundle)
    _write_pillar_reference(wb, bundle)
    _write_indicator_reference(wb, bundle)
    _write_method_guide(wb)
    wb.save(OUTPUT_XLSX)

    logger.info("Done in %.1fs", _time.time() - t0)
    logger.info("  Countries in bundle: %d", len(bundle["countries"]))
    logger.info("  Indicators documented: %d", len(bundle["indicators"]))
    logger.info("  Output: %s  +  %s", OUTPUT_JSON, OUTPUT_XLSX)
    logger.info("  Next step: python 07_dashboard.py")


if __name__ == "__main__":
    main()
