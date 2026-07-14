"""
05_robustness.py — Robustness and sensitivity analysis for AII scores.

Input:  data/04_scores.xlsx  (overall_scores, pillar_scores_wide sheets)
        data/02_clean.xlsx   (clean_data, fill_log sheets)
        data/03_norm.xlsx    (norm_data, norm_bounds sheets)
Output: data/05_robustness.xlsx

Analyses:
  1. Rank correlation matrix    — Spearman rho between all 5 scoring methods.
                                   High rho (>0.90) = rankings are robust.
  2. Country rank sensitivity   — Per country: min/max/range of ranks across methods.
                                   Large range = method-sensitive country.
  3. MaxS weight search         — random-restart grid search finds the weight vector
                                   that produces rankings MOST different from equal-weight.
                                   Worst-case robustness test (Saisana et al. 2005).
  4. Fill exclusion test        — Re-score using only real (non-filled) indicator values.
                                   Ranks countries should be stable; large shifts = data-quality risk.
  5. BoD bound sensitivity      — Re-run BoD with tighter bounds [0.07, 0.20] vs default
                                   [0.05, 0.25]. Large rank shift = BoD result is bound-sensitive.
  6. Robustness summary         — Overall verdict per country and per method.

Methodological grounding:
  Saisana, M., Saltelli, A., & Tarantola, S. (2005). Journal of the Royal
  Statistical Society: Series A, 168(2), 307-323. DOI: 10.1111/j.1467-985X.2005.00359.x

Run from the project root:
    python 05_robustness.py
"""

import sys
import logging
import time as _time
import json
import warnings

import numpy as np
import pandas as pd
import yaml
from pathlib import Path
from scipy.stats import spearmanr
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

from constants import PILLAR_DEFS, WEIGHT_MIN, WEIGHT_MAX, SMALL, ISLAND_SET

# ── Logging ────────────────────────────────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format="%(levelname)-8s %(message)s",
    stream=sys.stdout,
)
logger = logging.getLogger(__name__)

# ── Paths ──────────────────────────────────────────────────────────────────────

SCORES_FILE  = Path("data/04_scores.xlsx")
CLEAN_FILE   = Path("data/02_clean.xlsx")
NORM_FILE    = Path("data/03_norm.xlsx")
INDICATORS_DIR = Path("indicators_list")
OUTPUT_DIR   = Path("data")
OUTPUT_FILE  = OUTPUT_DIR / "05_robustness.xlsx"
OUTPUT_JSON  = OUTPUT_DIR / "05_robustness.json"

# ── Config ─────────────────────────────────────────────────────────────────────
# PILLAR_DEFS, WEIGHT_MIN, WEIGHT_MAX, SMALL, ISLAND_SET imported from constants.py

METHODS      = ["equal", "pca", "bod", "entropy", "geometric"]
BOD_ALT_MIN  = 0.07   # tighter bounds for BoD sensitivity test (vs default 0.05)
BOD_ALT_MAX  = 0.20   # tighter upper bound (vs default 0.25)

# ── Formatting helpers (shared style with 04_score.py) ─────────────────────────

HEADER_FILL = PatternFill(start_color="1B3A6B", end_color="1B3A6B", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT        = Alignment(horizontal="left",   vertical="center")


def _style_header(ws, n_cols: int):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
    ws.row_dimensions[1].height = 28


def _col_width(ws, col: int, width: float):
    ws.column_dimensions[get_column_letter(col)].width = width


def _color_scale(ws, col: int, min_row: int, max_row: int, low_good: bool = False):
    col_letter = get_column_letter(col)
    ref = f"{col_letter}{min_row}:{col_letter}{max_row}"
    if low_good:   # low value = green (e.g. rank range — smaller is better)
        rule = ColorScaleRule(
            start_type="min", start_color="64C882",
            mid_type="percentile", mid_value=50, mid_color="FFE664",
            end_type="max", end_color="F26464",
        )
    else:          # high value = green (e.g. correlation)
        rule = ColorScaleRule(
            start_type="min", start_color="F26464",
            mid_type="percentile", mid_value=50, mid_color="FFE664",
            end_type="max", end_color="64C882",
        )
    ws.conditional_formatting.add(ref, rule)


# ── 1. Rank correlation matrix ─────────────────────────────────────────────────

def rank_correlation_matrix(overall: pd.DataFrame) -> pd.DataFrame:
    """Spearman rho between every pair of scoring methods."""
    rank_cols = [f"{m}_rank" for m in METHODS if f"{m}_rank" in overall.columns]
    rank_df   = overall[rank_cols].dropna()
    n = len(METHODS)
    rho_mat = np.ones((n, n))
    for i, m1 in enumerate(METHODS):
        for j, m2 in enumerate(METHODS):
            if i == j:
                continue
            c1, c2 = f"{m1}_rank", f"{m2}_rank"
            if c1 in rank_df.columns and c2 in rank_df.columns:
                rho, _ = spearmanr(rank_df[c1], rank_df[c2])
                rho_mat[i, j] = round(float(rho), 4)
    return pd.DataFrame(rho_mat, index=METHODS, columns=METHODS)


# ── 2. Country rank sensitivity ────────────────────────────────────────────────

def country_rank_sensitivity(overall: pd.DataFrame, name_map: dict, region_map: dict,
                              island_set: set) -> pd.DataFrame:
    """Per-country min/max/range of ranks across all 5 methods."""
    rank_cols = [f"{m}_rank" for m in METHODS if f"{m}_rank" in overall.columns]
    rows = []
    for _, row in overall.iterrows():
        iso3 = row["iso3"]
        ranks = [row[c] for c in rank_cols if pd.notna(row.get(c))]
        if not ranks:
            continue
        min_r, max_r = int(min(ranks)), int(max(ranks))
        rows.append({
            "iso3":        iso3,
            "country":     name_map.get(iso3, iso3),
            "region":      region_map.get(iso3, ""),
            "island":      "Yes" if iso3 in island_set else "",
            **{f"{m}_rank": int(row[f"{m}_rank"]) if pd.notna(row.get(f"{m}_rank")) else None
               for m in METHODS if f"{m}_rank" in overall.columns},
            "rank_min":    min_r,
            "rank_max":    max_r,
            "rank_range":  max_r - min_r,
            "verdict":     "Stable" if (max_r - min_r) <= 5
                           else "Moderate" if (max_r - min_r) <= 12
                           else "Sensitive",
        })
    return pd.DataFrame(rows).sort_values("rank_range", ascending=False)


# ── 3. MaxS weight search ──────────────────────────────────────────────────────

def maxs_search(pillar_wide: pd.DataFrame) -> dict:
    """
    Find the weight vector that produces rankings MOST different from equal-weight.
    Minimises Spearman rho between equal-weight ranks and alternative-weight ranks.

    Implementation: random-restart grid search over the feasible weight simplex.
    SLSQP (the original approach) is not used because Spearman rank correlation is
    a step function in weight space — gradient-based optimisers converge trivially
    to the starting point without exploring. A grid search with deterministic corner
    cases (one pillar maximally weighted) correctly identifies the most rank-disrupting
    weight allocation.

    Decision (OECD Handbook §6.3; Saisana, Saltelli & Tarantola 2005, JRSS-A 168:307):
    The MaxS test is a worst-case robustness check — it asks "how bad could rankings
    get if a biased analyst chose the most adversarial weights within reasonable bounds?"
    1000 random starts from a Dirichlet(1,...,1) prior + 7 deterministic corners gives
    good coverage of the feasible simplex without gradient assumptions.
    """
    pillars = [p for p in PILLAR_DEFS if p in pillar_wide.columns]
    n = len(pillars)

    mat = pillar_wide[pillars].values.astype(float)
    valid_rows = ~np.isnan(mat).all(axis=1)
    mat = mat[valid_rows]

    equal_scores = np.nanmean(mat, axis=1)
    equal_ranks  = pd.Series(equal_scores).rank(ascending=False).values

    def _weighted_scores(w_arr):
        result = []
        for row in mat:
            valid = ~np.isnan(row)
            if not valid.any():
                result.append(0.0)
                continue
            w_v = w_arr[valid] / w_arr[valid].sum()
            result.append(float(np.dot(w_v, row[valid])))
        return np.array(result)

    def _rho(w_arr):
        alt_ranks = pd.Series(_weighted_scores(w_arr)).rank(ascending=False).values
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            rho, _ = spearmanr(equal_ranks, alt_ranks)
        return float(rho)

    rng      = np.random.default_rng(42)
    best_rho = float("inf")
    best_w   = np.ones(n) / n

    # Deterministic corners: concentrate maximum weight on each pillar in turn.
    # With WEIGHT_MAX=0.25 and n=7, remaining pillars get (1-0.25)/6 ≈ 0.125 each.
    for i in range(n):
        other_w = (1.0 - WEIGHT_MAX) / (n - 1)
        w = np.full(n, other_w)
        w[i] = WEIGHT_MAX
        if np.all(w >= WEIGHT_MIN - 1e-9) and np.all(w <= WEIGHT_MAX + 1e-9):
            rho = _rho(w)
            if rho < best_rho:
                best_rho, best_w = rho, w.copy()

    # Random restart: 1000 Dirichlet(1,...,1) samples, filtered to bounds.
    N_STARTS = 1000
    accepted = 0
    for _ in range(N_STARTS):
        raw = rng.dirichlet(np.ones(n))
        if np.all(raw >= WEIGHT_MIN) and np.all(raw <= WEIGHT_MAX):
            rho = _rho(raw)
            accepted += 1
            if rho < best_rho:
                best_rho, best_w = rho, raw.copy()

    logger.info("  MaxS: %d/%d random samples within bounds; best rho=%.4f",
                accepted, N_STARTS, best_rho)

    w_opt    = best_w
    rho_maxs = best_rho

    scores_maxs = _weighted_scores(w_opt)
    maxs_ranks  = pd.Series(scores_maxs).rank(ascending=False).values
    rank_shift  = np.abs(maxs_ranks - equal_ranks)

    return {
        "success":           True,
        "method":            "random_restart_grid_search",
        "n_starts":          N_STARTS,
        "weights":           {p: round(float(w), 4) for p, w in zip(pillars, w_opt)},
        "spearman_vs_equal": round(float(rho_maxs), 4),
        "mean_rank_shift":   round(float(np.mean(rank_shift)), 2),
        "max_rank_shift":    round(float(np.max(rank_shift)), 1),
        "verdict":           (
            "Highly robust" if rho_maxs >= 0.95 else
            "Robust"        if rho_maxs >= 0.85 else
            "Moderate"      if rho_maxs >= 0.70 else
            "Sensitive"
        ),
    }


# ── 4. Fill exclusion test ─────────────────────────────────────────────────────

def fill_exclusion_test(norm_df: pd.DataFrame, fill_log: pd.DataFrame,
                        pillar_map: dict) -> pd.DataFrame:
    """
    Re-score using only real (non-filled) indicator values.
    Sets normalized scores to NaN for any (iso3, variable_name) that was filled.
    Re-computes pillar means and equal-weight overall score.
    Returns comparison of ranks: main vs fill-excluded.
    """
    # Pivot norm scores to wide
    norm_wide = norm_df.pivot(index="iso3", columns="variable_name", values="norm_score").copy()
    norm_wide.columns.name = None

    # Mask filled observations
    n_masked = 0
    for _, row in fill_log.iterrows():
        iso3 = row.get("iso3")
        var  = row.get("variable_name")
        if iso3 in norm_wide.index and var in norm_wide.columns:
            norm_wide.loc[iso3, var] = float("nan")
            n_masked += 1
    logger.info("  Fill exclusion: masked %d filled observations", n_masked)

    # Recompute pillar scores
    pillar_scores = {}
    for p, vars_in in pillar_map.items():
        available = [v for v in vars_in if v in norm_wide.columns]
        if not available:
            continue
        sub = norm_wide[available]
        col = sub.mean(axis=1, skipna=True)
        col[sub.isna().all(axis=1)] = float("nan")
        pillar_scores[p] = col

    pillar_df = pd.DataFrame(pillar_scores, index=norm_wide.index)
    pillars   = [p for p in PILLAR_DEFS if p in pillar_df.columns]
    excl_score = pillar_df[pillars].mean(axis=1, skipna=True)
    excl_rank  = excl_score.rank(ascending=False, method="min", na_option="bottom")

    return excl_rank.rename("excl_rank"), excl_score.rename("excl_score")


# ── 5. BoD bound sensitivity ───────────────────────────────────────────────────

def bod_bound_sensitivity(pillar_wide: pd.DataFrame,
                          alt_min: float, alt_max: float) -> pd.Series:
    """Re-run BoD with alternative weight bounds and return ranks."""
    try:
        import pulp
    except ImportError:
        logger.warning("  BoD sensitivity skipped — pulp not installed")
        return pd.Series(dtype=float)

    pillars   = [p for p in PILLAR_DEFS if p in pillar_wide.columns]
    n_pillars = len(pillars)
    sub       = pillar_wide[pillars] / 100.0

    scores = {}
    for iso3, row in sub.iterrows():
        vals = row.values.astype(float)
        if np.isnan(vals).all():
            scores[iso3] = float("nan")
            continue
        prob = pulp.LpProblem(f"BoD_alt_{iso3}", pulp.LpMaximize)
        w = [pulp.LpVariable(f"w{i}", lowBound=alt_min, upBound=alt_max)
             for i in range(n_pillars)]
        prob += pulp.lpSum(w) == 1.0
        own = pulp.lpSum(w[i] * float(vals[i]) for i in range(n_pillars)
                         if not np.isnan(vals[i]))
        prob += own
        for other_iso3, other_row in sub.iterrows():
            if other_iso3 == iso3:
                continue
            ov = other_row.values.astype(float)
            vi = [i for i in range(n_pillars) if not np.isnan(ov[i])]
            if vi:
                prob += pulp.lpSum(w[i] * float(ov[i]) for i in vi) <= 1.0
        prob.solve(pulp.PULP_CBC_CMD(msg=0))
        scores[iso3] = float(own.value()) * 100.0 if pulp.LpStatus[prob.status] == "Optimal" \
                       else float("nan")

    s = pd.Series(scores)
    return s.rank(ascending=False, method="min", na_option="bottom").rename("bod_alt_rank")


# ── Output writers ─────────────────────────────────────────────────────────────

def _write_corr_sheet(wb, corr_df: pd.DataFrame):
    ws = wb.create_sheet("Rank Correlations")
    methods = list(corr_df.columns)
    # Header
    ws.cell(row=1, column=1, value="Method").fill = HEADER_FILL
    ws.cell(row=1, column=1).font = HEADER_FONT
    for c, m in enumerate(methods, 2):
        cell = ws.cell(row=1, column=c, value=m)
        cell.fill = HEADER_FILL; cell.font = HEADER_FONT; cell.alignment = CENTER
    _col_width(ws, 1, 14)
    for r, m in enumerate(methods, 2):
        ws.cell(row=r, column=1, value=m).font = Font(bold=True, size=10)
        for c, val in enumerate(corr_df.loc[m], 2):
            cell = ws.cell(row=r, column=c, value=round(float(val), 4))
            cell.number_format = "0.0000"
            cell.alignment = CENTER
            if r == c:
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        _col_width(ws, r, 14)
    # Color scale on data cells
    n = len(methods)
    for c in range(2, n + 2):
        col_letter = get_column_letter(c)
        ws.conditional_formatting.add(
            f"{col_letter}2:{col_letter}{n+1}",
            ColorScaleRule(start_type="num", start_value=0.5, start_color="F26464",
                           mid_type="num",   mid_value=0.9,   mid_color="FFE664",
                           end_type="num",   end_value=1.0,   end_color="64C882")
        )
    ws.freeze_panes = "B2"

    # Interpretation note
    note_row = n + 3
    ws.cell(row=note_row, column=1,
            value="Interpretation: rho > 0.90 = highly consistent rankings across methods. "
                  "rho < 0.80 = rankings are sensitive to method choice.").font = Font(italic=True, size=9)


def _write_sensitivity_sheet(wb, sens_df: pd.DataFrame):
    ws = wb.create_sheet("Country Sensitivity")
    headers = list(sens_df.columns)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h.replace("_", " ").title())
        cell.fill = HEADER_FILL; cell.font = HEADER_FONT; cell.alignment = CENTER
    ws.row_dimensions[1].height = 28

    verdict_colors = {"Stable": "eefaf4", "Moderate": "fff8e1", "Sensitive": "fdecea"}
    range_col = headers.index("rank_range") + 1

    for r, (_, row) in enumerate(sens_df.iterrows(), 2):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val if pd.notna(val) else None)
            if headers[c-1] == "verdict":
                cell.fill = PatternFill(
                    start_color=verdict_colors.get(str(val), "FFFFFF"),
                    end_color=verdict_colors.get(str(val), "FFFFFF"),
                    fill_type="solid"
                )
                cell.alignment = CENTER
            elif headers[c-1] in ("rank_range", "rank_min", "rank_max") or "rank" in headers[c-1]:
                cell.number_format = "0"
                cell.alignment = CENTER
            else:
                cell.alignment = LEFT

    _color_scale(ws, range_col, 2, len(sens_df) + 1, low_good=True)

    widths = [7, 22, 10, 6] + [10] * len(METHODS) + [10, 10, 12, 12]
    for c, w in enumerate(widths, 1):
        _col_width(ws, c, w)

    ws.freeze_panes = "E2"
    ws.auto_filter.ref = ws.dimensions

    ws.cell(row=len(sens_df) + 3, column=1,
            value="Rank Range = max rank - min rank across 5 methods. "
                  "Stable ≤5 | Moderate ≤12 | Sensitive >12").font = Font(italic=True, size=9)


def _write_maxs_sheet(wb, maxs: dict, pillar_wide: pd.DataFrame):
    ws = wb.create_sheet("MaxS Weight Search")
    ws.cell(row=1, column=1, value="MaxS Robustness Test").fill = HEADER_FILL
    ws.cell(row=1, column=1).font = HEADER_FONT

    info = [
        ("Result", "Success" if maxs["success"] else "Optimisation failed"),
        ("Spearman vs Equal-Weight", f"{maxs['spearman_vs_equal']:.4f}"),
        ("Mean Rank Shift", f"{maxs['mean_rank_shift']:.1f} positions"),
        ("Max Rank Shift", f"{maxs['max_rank_shift']:.0f} positions"),
        ("Overall Verdict", maxs["verdict"]),
    ]
    for r, (label, val) in enumerate(info, 2):
        ws.cell(row=r, column=1, value=label).font = Font(bold=True, size=10)
        ws.cell(row=r, column=2, value=val)

    ws.cell(row=8, column=1, value="Worst-Case Weights").font = Font(bold=True, size=10)
    ws.cell(row=9, column=1, value="Pillar").fill = HEADER_FILL
    ws.cell(row=9, column=1).font = HEADER_FONT
    ws.cell(row=9, column=2, value="MaxS Weight").fill = HEADER_FILL
    ws.cell(row=9, column=2).font = HEADER_FONT
    ws.cell(row=9, column=3, value="Equal Weight").fill = HEADER_FILL
    ws.cell(row=9, column=3).font = HEADER_FONT
    ws.cell(row=9, column=4, value="Difference").fill = HEADER_FILL
    ws.cell(row=9, column=4).font = HEADER_FONT

    for r, (pillar_id, pillar_name) in enumerate(PILLAR_DEFS.items(), 10):
        maxs_w = maxs["weights"].get(pillar_id, float("nan"))
        equal_w = round(1/7, 4)
        ws.cell(row=r, column=1, value=f"{pillar_id} — {pillar_name}")
        ws.cell(row=r, column=2, value=maxs_w).number_format = "0.0000"
        ws.cell(row=r, column=3, value=equal_w).number_format = "0.0000"
        diff = maxs_w - equal_w if isinstance(maxs_w, float) else None
        cell = ws.cell(row=r, column=4, value=round(diff, 4) if diff is not None else None)
        cell.number_format = "+0.0000;-0.0000"
        if diff and diff > 0:
            cell.font = Font(color="27ae60", bold=True)
        elif diff and diff < 0:
            cell.font = Font(color="c0392b", bold=True)

    note = (
        f"Interpretation: Even in the worst case (rho={maxs['spearman_vs_equal']:.3f}), "
        f"the mean rank shift is {maxs['mean_rank_shift']:.1f} positions. "
        f"Verdict: {maxs['verdict']}."
    )
    ws.cell(row=19, column=1, value=note).font = Font(italic=True, size=9)

    _col_width(ws, 1, 32); _col_width(ws, 2, 14); _col_width(ws, 3, 14); _col_width(ws, 4, 14)


def _write_fill_exclusion_sheet(wb, overall: pd.DataFrame,
                                 excl_rank: pd.Series, excl_score: pd.Series,
                                 name_map: dict, island_set: set):
    ws = wb.create_sheet("Fill Exclusion Test")
    headers = ["ISO3", "Country", "Island", "Equal Rank\n(Main)", "Equal Rank\n(Real Data Only)",
               "Rank Shift", "Verdict"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL; cell.font = HEADER_FONT; cell.alignment = CENTER
    ws.row_dimensions[1].height = 36

    main_ranks = overall.set_index("iso3")["equal_rank"]

    rows = []
    for iso3 in overall["iso3"]:
        main_r = main_ranks.get(iso3)
        excl_r = excl_rank.get(iso3)
        shift  = int(abs(excl_r - main_r)) if pd.notna(excl_r) and pd.notna(main_r) else None
        rows.append({
            "iso3": iso3, "country": name_map.get(iso3, iso3),
            "island": "Yes" if iso3 in island_set else "",
            "main_rank": int(main_r) if pd.notna(main_r) else None,
            "excl_rank": int(excl_r) if pd.notna(excl_r) else None,
            "shift": shift,
            "verdict": ("Stable" if shift is not None and shift <= 3
                        else "Moderate" if shift is not None and shift <= 8
                        else "Data-sensitive"),
        })

    rows_df = pd.DataFrame(rows).sort_values("shift", ascending=False, na_position="last")
    verdict_colors = {"Stable": "eefaf4", "Moderate": "fff8e1", "Data-sensitive": "fdecea"}

    for r, (_, row) in enumerate(rows_df.iterrows(), 2):
        vals = [row["iso3"], row["country"], row["island"], row["main_rank"],
                row["excl_rank"], row["shift"], row["verdict"]]
        for c, val in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=val if pd.notna(val) else None)
            if headers[c-1] in ("Equal Rank\n(Main)", "Equal Rank\n(Real Data Only)", "Rank Shift"):
                cell.number_format = "0"; cell.alignment = CENTER
            elif headers[c-1] == "Verdict":
                cell.fill = PatternFill(
                    start_color=verdict_colors.get(str(val), "FFFFFF"),
                    end_color=verdict_colors.get(str(val), "FFFFFF"), fill_type="solid")
                cell.alignment = CENTER
            else:
                cell.alignment = LEFT

    shift_col = 6
    _color_scale(ws, shift_col, 2, len(rows_df) + 1, low_good=True)
    widths = [7, 22, 7, 14, 18, 12, 16]
    for c, w in enumerate(widths, 1):
        _col_width(ws, c, w)
    ws.freeze_panes = "D2"
    ws.auto_filter.ref = ws.dimensions

    ws.cell(row=len(rows_df) + 3, column=1,
            value="Rank Shift = |main rank - rank using real data only|. "
                  "Stable ≤3 | Moderate ≤8 | Data-sensitive >8").font = Font(italic=True, size=9)


def _write_bod_sensitivity_sheet(wb, overall: pd.DataFrame,
                                  alt_rank: pd.Series, name_map: dict):
    ws = wb.create_sheet("BoD Bound Sensitivity")
    headers = ["ISO3", "Country",
               f"BoD Rank\n[{WEIGHT_MIN},{WEIGHT_MAX}]",
               f"BoD Rank\n[{BOD_ALT_MIN},{BOD_ALT_MAX}]",
               "Rank Shift"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL; cell.font = HEADER_FONT; cell.alignment = CENTER
    ws.row_dimensions[1].height = 36

    main_bod = overall.set_index("iso3")["bod_rank"]
    rows = []
    for iso3 in overall["iso3"]:
        main_r = main_bod.get(iso3)
        alt_r  = alt_rank.get(iso3)
        shift  = int(abs(alt_r - main_r)) if pd.notna(alt_r) and pd.notna(main_r) else None
        rows.append({"iso3": iso3, "country": name_map.get(iso3, iso3),
                     "main": int(main_r) if pd.notna(main_r) else None,
                     "alt":  int(alt_r)  if pd.notna(alt_r)  else None,
                     "shift": shift})

    rows_df = pd.DataFrame(rows).sort_values("shift", ascending=False, na_position="last")
    for r, (_, row) in enumerate(rows_df.iterrows(), 2):
        vals = [row["iso3"], row["country"], row["main"], row["alt"], row["shift"]]
        for c, val in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=val if pd.notna(val) else None)
            if c >= 3:
                cell.number_format = "0"; cell.alignment = CENTER
            else:
                cell.alignment = LEFT

    _color_scale(ws, 5, 2, len(rows_df) + 1, low_good=True)
    for c, w in enumerate([7, 22, 16, 16, 12], 1):
        _col_width(ws, c, w)
    ws.auto_filter.ref = ws.dimensions


def _write_summary_sheet(wb, sens_df: pd.DataFrame, corr_df: pd.DataFrame,
                          maxs: dict, name_map: dict):
    ws = wb.create_sheet("Robustness Summary", 0)  # first sheet
    ws.cell(row=1, column=1, value="AII Robustness Summary").fill = HEADER_FILL
    ws.cell(row=1, column=1).font = Font(bold=True, color="FFFFFF", size=13)
    ws.merge_cells("A1:D1")

    # Overall stats
    min_rho = corr_df.values[np.triu_indices_from(corr_df.values, k=1)].min()
    stable   = (sens_df["verdict"] == "Stable").sum()
    moderate = (sens_df["verdict"] == "Moderate").sum()
    sensitive= (sens_df["verdict"] == "Sensitive").sum()

    stats = [
        ("Min pairwise rank correlation", f"{min_rho:.4f}",
         "High" if min_rho >= 0.90 else "Moderate" if min_rho >= 0.80 else "Low"),
        ("Countries: stable across methods", str(stable), ""),
        ("Countries: moderate sensitivity", str(moderate), ""),
        ("Countries: high sensitivity", str(sensitive),
         "Review" if sensitive > 5 else ""),
        ("MaxS worst-case verdict", maxs["verdict"], ""),
        ("MaxS mean rank shift", f"{maxs['mean_rank_shift']:.1f} positions", ""),
    ]
    for r, (label, val, flag) in enumerate(stats, 3):
        ws.cell(row=r, column=1, value=label).font = Font(bold=True, size=10)
        ws.cell(row=r, column=2, value=val)
        if flag:
            cell = ws.cell(row=r, column=3, value=flag)
            color = "eefaf4" if flag in ("High",) else "fff8e1" if flag in ("Moderate",) else "fdecea"
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    # Most sensitive countries
    ws.cell(row=11, column=1, value="Most Method-Sensitive Countries").font = Font(bold=True, size=10)
    top_sensitive = sens_df[sens_df["verdict"] == "Sensitive"].head(10)
    for r, (_, row) in enumerate(top_sensitive.iterrows(), 12):
        ws.cell(row=r, column=1, value=name_map.get(row["iso3"], row["iso3"]))
        ws.cell(row=r, column=2, value=f"Rank range: {int(row['rank_range'])} positions")
        ws.cell(row=r, column=3, value=row["verdict"]).fill = PatternFill(
            start_color="fdecea", end_color="fdecea", fill_type="solid")

    _col_width(ws, 1, 36); _col_width(ws, 2, 28); _col_width(ws, 3, 14)


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    t0 = _time.time()

    for f in [SCORES_FILE, CLEAN_FILE, NORM_FILE]:
        if not f.exists():
            logger.error("Required file not found: %s", f)
            sys.exit(1)

    logger.info("Step 1: Load scores and pillar data")
    overall     = pd.read_excel(SCORES_FILE, sheet_name="Overall Rankings")
    pillar_wide = pd.read_excel(SCORES_FILE, sheet_name="Pillar Scores")

    # Normalise column names from the formatted sheet
    overall.columns = [c.replace("\n", " ").strip() for c in overall.columns]

    # Rebuild iso3 column (Overall Rankings has "ISO3" not "iso3")
    if "ISO3" in overall.columns:
        overall = overall.rename(columns={"ISO3": "iso3"})

    # Rebuild method rank columns (sheet has "Equal Weights Rank" etc — map back)
    col_map = {
        "Equal Weights Score":    "equal",
        "PCA Weights Score":      "pca",
        "Benefit of Doubt Score": "bod",
        "Entropy Weights Score":  "entropy",
        "Geometric Mean Score":   "geometric",
        "Equal Weights Rank":     "equal_rank",
        "PCA Weights Rank":       "pca_rank",
        "Benefit of Doubt Rank":  "bod_rank",
        "Entropy Weights Rank":   "entropy_rank",
        "Geometric Mean Rank":    "geometric_rank",
    }
    overall = overall.rename(columns=col_map)

    # Pillar wide sheet
    pillar_wide.columns = [c.replace("\n", " ").strip() for c in pillar_wide.columns]
    if "ISO3" in pillar_wide.columns:
        pillar_wide = pillar_wide.rename(columns={"ISO3": "iso3"})
    pillar_wide = pillar_wide.set_index("iso3")
    # Map full pillar name columns (e.g. "Political / Governance Score") back to letters
    full_to_letter = {f"{v} Score": k for k, v in PILLAR_DEFS.items()}
    pillar_score_cols = [c for c in pillar_wide.columns if c in full_to_letter]
    pillar_wide = pillar_wide[pillar_score_cols].rename(columns=full_to_letter)

    logger.info("Step 2: Load fill log and norm data")
    fill_log = pd.read_excel(CLEAN_FILE, sheet_name="fill_log")
    norm_df  = pd.read_excel(NORM_FILE,  sheet_name="norm_data")

    # Load pillar map
    pillar_map: dict[str, list[str]] = {k: [] for k in PILLAR_DEFS}
    for path in sorted(INDICATORS_DIR.glob("*.yaml")):
        with open(path, "r", encoding="utf-8") as f:
            data = yaml.safe_load(f)
        if not data:
            continue
        for ind in data:
            if ind.get("role", "scoring") != "scoring":
                continue
            for p in ind.get("pillars", []):
                if p in pillar_map and ind["variable_name"] not in pillar_map[p]:
                    pillar_map[p].append(ind["variable_name"])

    from models.countries import COUNTRIES
    name_map   = {k: v["name"]   for k, v in COUNTRIES.items()}
    region_map = {k: v["region"] for k, v in COUNTRIES.items()}
    # ISLAND_SET imported from constants.py — do not redefine here

    logger.info("Step 3: Rank correlation matrix")
    corr_df = rank_correlation_matrix(overall)
    min_rho = corr_df.values[np.triu_indices_from(corr_df.values, k=1)].min()
    logger.info("  Min pairwise Spearman rho: %.4f", min_rho)

    logger.info("Step 4: Country rank sensitivity")
    sens_df = country_rank_sensitivity(overall, name_map, region_map, ISLAND_SET)
    stable   = (sens_df["verdict"] == "Stable").sum()
    sensitive = (sens_df["verdict"] == "Sensitive").sum()
    logger.info("  Stable: %d  |  Sensitive: %d", stable, sensitive)

    logger.info("Step 5: MaxS weight search (random-restart grid search)")
    maxs = maxs_search(pillar_wide)
    logger.info("  Verdict: %s  |  Spearman vs equal: %.4f  |  Max rank shift: %.0f",
                maxs["verdict"], maxs["spearman_vs_equal"], maxs["max_rank_shift"])

    logger.info("Step 6: Fill exclusion test")
    excl_rank, excl_score = fill_exclusion_test(norm_df, fill_log, pillar_map)

    logger.info("Step 7: BoD bound sensitivity [%.2f, %.2f]", BOD_ALT_MIN, BOD_ALT_MAX)
    alt_bod_rank = bod_bound_sensitivity(pillar_wide, BOD_ALT_MIN, BOD_ALT_MAX)

    logger.info("Step 8: Write output to %s", OUTPUT_FILE)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        # Write a blank sheet first so we can insert Summary as sheet 0 later
        pd.DataFrame().to_excel(writer, sheet_name="_placeholder", index=False)

    from openpyxl import load_workbook
    wb = load_workbook(OUTPUT_FILE)
    wb.remove(wb["_placeholder"])

    _write_summary_sheet(wb, sens_df, corr_df, maxs, name_map)
    _write_corr_sheet(wb, corr_df)
    _write_sensitivity_sheet(wb, sens_df)
    _write_maxs_sheet(wb, maxs, pillar_wide)
    _write_fill_exclusion_sheet(wb, overall, excl_rank, excl_score, name_map, ISLAND_SET)
    if not alt_bod_rank.empty:
        _write_bod_sensitivity_sheet(wb, overall, alt_bod_rank, name_map)
    wb.save(OUTPUT_FILE)

    # JSON summary
    summary = {
        "min_rank_correlation":   round(float(min_rho), 4),
        "n_stable":               int(stable),
        "n_sensitive":            int(sensitive),
        "maxs":                   maxs,
        "rank_correlation_matrix": {
            m: {m2: round(float(corr_df.loc[m, m2]), 4) for m2 in METHODS}
            for m in METHODS
        },
    }
    with open(OUTPUT_JSON, "w") as f:
        json.dump(summary, f, indent=2)

    logger.info("Done in %.1fs", _time.time() - t0)
    logger.info("  Output: %s  +  %s", OUTPUT_FILE, OUTPUT_JSON)


if __name__ == "__main__":
    main()
